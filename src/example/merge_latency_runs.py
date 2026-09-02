#!/usr/bin/env python3
"""
merge_latency_runs.py

Merges all latency_run_*.csv files in ~/latency_data/ into a single Excel file.
Each run becomes one column. Rows are frame indices.

Usage:
    python3 merge_latency_runs.py
    python3 merge_latency_runs.py --input ~/latency_data --output ~/latency_results.xlsx
"""

import argparse
import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt


def parse_args():
    parser = argparse.ArgumentParser(description="Merge latency CSV runs into Excel")
    parser.add_argument("--input",  default=os.path.expanduser("~/latency_data/native_iceoryx/"),
                        help="Directory containing latency_run_*.csv files")
    parser.add_argument("--output", default=os.path.expanduser("~/latency_data/native_iceoryx/latency_results.xlsx"),
                        help="Output Excel file path")
    return parser.parse_args()


def load_runs(input_dir, i):
    """Load all CSV run files, return dict of {run_label: Series of full_us values}"""
    pattern = os.path.join(input_dir, "*latency_run_*.csv")
    files = sorted(glob.glob(pattern))

    if not files:
        raise FileNotFoundError(f"No latency_run_*.csv files found in {input_dir}")

    print(i)
    if i != 3:
        runs = {
            "full_us": pd.Series(dtype=float),
            "transport_us": pd.Series(dtype=float)
                }
        for f in files:
            label = os.path.splitext(os.path.basename(f))[0]  # e.g. latency_run_2026-03-13_12-00-00
            df = pd.read_csv(f)
            if "full_us" not in df.columns:
                print(f"[WARN] Skipping {f} — no 'full_us' column")
                continue
            df["full_us"] = df["full_us"]/1000
            df["transport_us"] = df["transport_us"]/1000
            runs["full_us"] = pd.concat([runs["full_us"], df["full_us"].reset_index(drop=True)], ignore_index=True) 
            runs["transport_us"] = pd.concat([runs["transport_us"], df["transport_us"].reset_index(drop=True)], ignore_index=True) 
            print(f"[INFO] Loaded {label}: {len(runs['full_us'])} frames")
            print(f"[INFO] Loaded {label}: {len(runs['transport_us'])} frames")
    else:
        runs = {
            "full_ms": pd.Series(dtype=float),
            "transport_ms": pd.Series(dtype=float)
                }
        for f in files:
            label = os.path.splitext(os.path.basename(f))[0]  # e.g. latency_run_2026-03-13_12-00-00
            df = pd.read_csv(f)
            if "full_ms" not in df.columns:
                print(f"[WARN] Skipping {f} — no 'full_ms' column")
                continue
            df["full_ms"] = df["full_ms"]
            df["transport_ms"] = df["transport_ms"]
            runs["full_ms"] = pd.concat([runs["full_ms"], df["full_ms"].reset_index(drop=True)], ignore_index=True) 
            runs["transport_ms"] = pd.concat([runs["transport_ms"], df["transport_ms"].reset_index(drop=True)], ignore_index=True) 
            print(f"[INFO] Loaded {label}: {len(runs['full_ms'])} frames")
            print(f"[INFO] Loaded {label}: {len(runs['transport_ms'])} frames")

    return runs

'''
def load_runs(input_dir):
    """
    Load CSV run files and group them by prefix before '_latency_run_'.
    Returns dict {group_name: merged Series}
    """
    pattern = os.path.join(input_dir, "*_latency_run_*.csv")
    files = sorted(glob.glob(pattern))

    if not files:
        raise FileNotFoundError(f"No latency_run_*.csv files found in {input_dir}")

    grouped = {}

    for f in files:
        filename = os.path.splitext(os.path.basename(f))[0]

        if "_latency_run_" not in filename:
            print(f"[WARN] Skipping {f} — invalid naming")
            continue

        # extract group name (prefix before _latency_run_)
        group_name = filename.split("_latency_run_")[0]

        df = pd.read_csv(f)
        if "full_us" not in df.columns:
            print(f"[WARN] Skipping {f} — no 'full_us' column")
            continue

        series = df["full_us"].dropna().reset_index(drop=True)

        if group_name not in grouped:
            grouped[group_name] = []

        grouped[group_name].append(series)

        print(f"[INFO] Loaded {filename} → group '{group_name}' ({len(series)} frames)")

    # merge runs per group (concatenate)
    merged_runs = {}
    for group_name, series_list in grouped.items():
        merged = pd.concat(series_list, ignore_index=True)
        merged_runs[group_name] = merged
        print(f"[INFO] Merged group '{group_name}': {len(merged)} total frames")

    return merged_runs
'''

def build_dataframe(runs):
    """Build combined DataFrame — one column per run, rows are frame indices"""
    df = pd.DataFrame(runs)
    df.index.name = "frame"
    return df

def remove_outliers(series):
    """Remove outliers using IQR method"""
    s = series.dropna()

    Q1 = s.quantile(0.25)
    Q3 = s.quantile(0.75)
    IQR = Q3 - Q1

    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR

    filtered = s[(s >= lower_bound) & (s <= upper_bound)]

    return filtered

def add_stats_sheet(wb, runs_df):
    """Add a summary statistics sheet"""
    ws = wb.create_sheet("Statistics")

    headers = ["Run", "Frames", "Mean (ms)", "Median (ms)", "P(95)",
               "Min (ms)", "Max (ms)", "Std Dev (ms)"]

    # header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    # data rows
    runs_df = runs_df.apply(remove_outliers)
    for row, run_name in enumerate(runs_df.columns, 2):
        s = runs_df[run_name]
        ws.cell(row=row, column=1, value=run_name).font = Font(name="Arial", bold=True)
        ws.cell(row=row, column=2, value=len(s))
        ws.cell(row=row, column=3, value=round(s.mean(), 3))
        ws.cell(row=row, column=4, value=round(s.median(), 3))
        ws.cell(row=row, column=5, value=round(s.quantile(0.95), 3))
        ws.cell(row=row, column=6, value=round(s.min(), 3))
        ws.cell(row=row, column=7, value=round(s.max(), 3))
        ws.cell(row=row, column=8, value=round(s.std(), 3))

        # alternate row shading
        fill_color = "D6E4F0" if row % 2 == 0 else "FFFFFF"
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.font = Font(name="Arial")
            c.alignment = Alignment(horizontal="center")

    # column widths
    col_widths = [40, 10, 12, 14, 10, 10, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    return ws


def add_raw_sheet(wb, runs_df):
    """Add raw latency data sheet — one column per run"""
    ws = wb.create_sheet("Raw Latency (µs)")

    # header
    ws.cell(row=1, column=1, value="Frame").font = Font(name="Arial", bold=True)
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for col, run_name in enumerate(runs_df.columns, 2):
        cell = ws.cell(row=1, column=col, value=run_name)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    # data rows
    for row_idx, frame_idx in enumerate(runs_df.index, 2):
        ws.cell(row=row_idx, column=1, value=frame_idx)
        for col_idx, run_name in enumerate(runs_df.columns, 2):
            val = runs_df.loc[frame_idx, run_name]
            if pd.notna(val):
                c = ws.cell(row=row_idx, column=col_idx, value=round(val, 2))
                c.font = Font(name="Arial")
                c.alignment = Alignment(horizontal="right")

    # column widths
    ws.column_dimensions["A"].width = 8
    for col in range(2, len(runs_df.columns) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 35

    return ws


def add_max_chart_sheet(wb, runs_df):
    """Add a line chart of latency over frames, excluding outliers."""

    ws = wb.create_sheet("Median Chart")

    # Remove outliers
    chart_df = runs_df.apply(remove_outliers)

    # Downsample ONLY for the chart
    # Change 100 to whatever gives the desired visual density
    BIN_SIZE = 50

    median_df = chart_df.groupby(chart_df.index // BIN_SIZE).median()
    max_df = chart_df.groupby(chart_df.index // BIN_SIZE).max()
    # Helper sheet containing only chart data
    data_ws = wb.create_sheet("Chart Data")

    data_ws.cell(row=1, column=1, value="Frame")

    for col_idx, run_name in enumerate(median_df.columns, 2):
        data_ws.cell(row=1, column=col_idx, value=run_name)

    for row_idx, frame_idx in enumerate(median_df.index, 2):
        data_ws.cell(row=row_idx, column=1, value=frame_idx)

        for col_idx, run_name in enumerate(median_df.columns, 2):
            value = median_df.loc[frame_idx, run_name]

            if pd.notna(value):
                data_ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=float(value)
                )

    data_ws.sheet_state = "hidden"

    # Chart
    chart = LineChart()
    chart.title = "Latency per Frame"
    chart.style = 10
    chart.y_axis.title = "Latency (ms)"
    chart.x_axis.title = "Frame"
    chart.width = 30
    chart.height = 15

    max_frames = len(median_df)

    for col_idx in range(2, len(median_df.columns) + 2):
        data_ref = Reference(
            data_ws,
            min_col=col_idx,
            min_row=1,
            max_row=max_frames + 1
        )

        chart.add_data(
            data_ref,
            titles_from_data=True
        )

    categories = Reference(
        data_ws,
        min_col=1,
        min_row=2,
        max_row=max_frames + 1
    )

    chart.set_categories(categories)

    ws.add_chart(chart, "A1")

    return ws

def add_chart_sheet(wb, runs_df):
    """Add a line chart of latency over frames, excluding outliers."""

    ws = wb.create_sheet("Max Chart")

    # Remove outliers
    chart_df = runs_df.apply(remove_outliers)

    # Downsample ONLY for the chart
    # Change 100 to whatever gives the desired visual density
    BIN_SIZE = 50

    max_df = chart_df.groupby(chart_df.index // BIN_SIZE).max()
    # Helper sheet containing only chart data
    data_ws = wb.create_sheet("Chart Max Data")

    data_ws.cell(row=1, column=1, value="Frame")

    for col_idx, run_name in enumerate(max_df.columns, 2):
        data_ws.cell(row=1, column=col_idx, value=run_name)

    for row_idx, frame_idx in enumerate(max_df.index, 2):
        data_ws.cell(row=row_idx, column=1, value=frame_idx)

        for col_idx, run_name in enumerate(max_df.columns, 2):
            value = max_df.loc[frame_idx, run_name]

            if pd.notna(value):
                data_ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=float(value)
                )

    data_ws.sheet_state = "hidden"

    # Chart
    chart = LineChart()
    chart.title = "Latency per Frame"
    chart.style = 10
    chart.y_axis.title = "Latency (ms)"
    chart.x_axis.title = "Frame"
    chart.width = 30
    chart.height = 15

    max_frames = len(max_df)

    for col_idx in range(2, len(max_df.columns) + 2):
        data_ref = Reference(
            data_ws,
            min_col=col_idx,
            min_row=1,
            max_row=max_frames + 1
        )

        chart.add_data(
            data_ref,
            titles_from_data=True
        )

    categories = Reference(
        data_ws,
        min_col=1,
        min_row=2,
        max_row=max_frames + 1
    )

    chart.set_categories(categories)

    ws.add_chart(chart, "A1")

    return ws

def main():
    args = parse_args()
    all_runs = []
    inputs = ["/home/alexandre/latency_data/ros2_iceoryx/", "/home/alexandre/latency_data/ros2_cyclonedds/", "/home/alexandre/latency_data/ros2_fastrtps/", "/home/alexandre/latency_data/native_iceoryx/"]
    outputs = ["/home/alexandre/latency_data/ros2_iceoryx/latency_results.xlsx", "/home/alexandre/latency_data/ros2_cyclonedds/latency_results.xlsx", "/home/alexandre/latency_data/ros2_fastrtps/latency_results.xlsx", "/home/alexandre/latency_data/native_iceoryx/latency_results.xlsx"]
    for i in range(len(inputs)):
        print(f"[INFO] Loading runs from {inputs[i]}")
        runs = load_runs(inputs[i], i)
        all_runs.append(runs)

        if not runs:
            print("[ERROR] No valid run data found")
            return

        runs_df = build_dataframe(runs)
        print(f"[INFO] Combined {len(runs)} runs, {len(runs_df)} max frames")

        # write to Excel via pandas first
        runs_df.to_excel(outputs[i], sheet_name="Raw Latency (µs)", index=True)

        # reopen with openpyxl for formatting
        wb = load_workbook(outputs[i])

        # remove default sheet — we'll recreate it formatted
        if "Raw Latency (µs)" in wb.sheetnames:
            del wb["Raw Latency (µs)"]

        # add sheets in order: Statistics first, then Raw, then Chart
        add_stats_sheet(wb, runs_df)
        add_raw_sheet(wb, runs_df)
        add_chart_sheet(wb, runs_df)
        add_max_chart_sheet(wb, runs_df)

        wb.save(outputs[i])
        print(f"[INFO] Saved {outputs[i]}")
        print(f"[INFO] Sheets: Statistics, Raw Latency (µs), Chart")

    fig_1 = plt.figure(figsize=(10, 7))
    ax_1 = fig_1.add_axes([0, 0, 1, 1])

    fig_2 = plt.figure(figsize=(10, 7))
    ax_2 = fig_2.add_axes([0, 0, 1, 1])

    boxplot_data = []
    boxplot_transport = []

    for i, runs in enumerate(all_runs):
        if i != 3:
            boxplot_data.append(runs["full_us"])
            boxplot_transport.append(runs["transport_us"])
        else:
            boxplot_data.append(runs["full_ms"])
            boxplot_transport.append(runs["transport_ms"])

    bp_1 = ax_1.boxplot(boxplot_data)
    bp_2 = ax_2.boxplot(boxplot_transport)

    ax_1.set_xticklabels([
        "ROS 2 Iceoryx",
        "ROS 2 CycloneDDS",
        "ROS 2 FastDDS",
        "Native Iceoryx"
    ])

    ax_2.set_xticklabels([
        "ROS 2 Iceoryx",
        "ROS 2 CycloneDDS",
        "ROS 2 FastDDS",
        "Native Iceoryx"
    ])

    ax_1.set_ylabel("Latency (ms)")
    ax_1.set_title("Full Latency Comparison")

    ax_2.set_ylabel("Latency (ms)")
    ax_2.set_title("Transport Latency Comparison")

    fig_1.savefig("/home/alexandre/Pictures/thesis/boxplot_full.png", dpi=300, bbox_inches="tight")
    fig_2.savefig("/home/alexandre/Pictures/thesis/boxplot_transport.png", dpi=300, bbox_inches="tight")

if __name__ == "__main__":
    main()
